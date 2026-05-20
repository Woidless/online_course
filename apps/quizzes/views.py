from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.http import HttpResponse

import csv
from io import StringIO

from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from apps.users.permissions import IsTeacherOrAdmin, IsStudent
from apps.courses.models import Enrollment
from apps.lessons.models import Lesson
from .models import Quiz, Question, Answer, QuizAttempt
from .serializers import (
    QuizSerializer,
    QuizAttemptSerializer,
    SubmitQuizSerializer,
    QuestionSerializer,
)


class QuizListCreateView(generics.ListCreateAPIView):
    """
    GET  /api/lessons/<lesson_id>/quizzes/ — список тестов урока
    POST /api/lessons/<lesson_id>/quizzes/ — создать тест (teacher, admin)
    """
    serializer_class = QuizSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        user = self.request.user
        if user.role in ('teacher', 'admin'):
            return Quiz.objects.filter(lesson_id=self.kwargs['lesson_id'])
        return Quiz.objects.filter(
            lesson_id=self.kwargs['lesson_id'],
            is_published=True
        )

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsTeacherOrAdmin()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        lesson = get_object_or_404(Lesson, pk=self.kwargs['lesson_id'])
        serializer.save(lesson=lesson)


class QuizDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    GET    /api/quizzes/<id>/ — детали теста
    PUT    /api/quizzes/<id>/ — редактировать (teacher, admin)
    DELETE /api/quizzes/<id>/ — удалить (teacher, admin)
    """
    serializer_class = QuizSerializer
    queryset = Quiz.objects.all()

    def get_permissions(self):
        if self.request.method == 'GET':
            return [IsAuthenticated()]
        return [IsTeacherOrAdmin()]


class QuestionCreateView(generics.CreateAPIView):
    """
    POST /api/quizzes/<quiz_id>/questions/ — добавить вопрос
    """
    serializer_class = QuestionSerializer
    permission_classes = (IsTeacherOrAdmin,)

    def perform_create(self, serializer):
        quiz = get_object_or_404(Quiz, pk=self.kwargs['quiz_id'])
        serializer.save(quiz=quiz)


class QuestionDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    PUT    /api/questions/<id>/ — редактировать вопрос
    DELETE /api/questions/<id>/ — удалить вопрос
    """
    serializer_class = QuestionSerializer
    permission_classes = (IsTeacherOrAdmin,)
    queryset = Question.objects.all()


class AnswerCreateView(APIView):
    """
    POST /api/questions/<question_id>/answers/ — добавить вариант ответа
    """
    permission_classes = (IsTeacherOrAdmin,)

    def post(self, request, question_id):
        question = get_object_or_404(Question, pk=question_id)
        answers_data = request.data.get('answers', [])

        if not isinstance(answers_data, list):
            return Response(
                {'detail': 'Ожидается список вариантов ответов.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Удаляем старые варианты и создаём новые
        question.answers.all().delete()
        created = []
        for item in answers_data:
            answer = Answer.objects.create(
                question=question,
                text=item.get('text', ''),
                is_correct=item.get('is_correct', False)
            )
            created.append({'id': answer.id, 'text': answer.text, 'is_correct': answer.is_correct})

        return Response(created, status=status.HTTP_201_CREATED)


class StartQuizView(APIView):
    """
    POST /api/quizzes/<quiz_id>/start/ — начать тест (student)
    """
    permission_classes = (IsStudent,)

    def post(self, request, quiz_id):
        quiz = get_object_or_404(Quiz, pk=quiz_id, is_published=True)

        # Проверяем запись на курс
        enrolled = Enrollment.objects.filter(
            student=request.user,
            group__course=quiz.lesson.course,
            status='active'
        ).exists()

        if not enrolled:
            return Response(
                {'detail': 'Вы не записаны на этот курс.'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Проверяем незавершённую попытку
        existing = QuizAttempt.objects.filter(
            quiz=quiz,
            student=request.user,
            status=QuizAttempt.Status.IN_PROGRESS
        ).first()

        if existing:
            return Response(
                QuizAttemptSerializer(existing).data,
                status=status.HTTP_200_OK
            )

        attempt = QuizAttempt.objects.create(
            quiz=quiz,
            student=request.user,
        )
        return Response(
            QuizAttemptSerializer(attempt).data,
            status=status.HTTP_201_CREATED
        )


class SubmitQuizView(APIView):
    """
    POST /api/attempts/<attempt_id>/submit/ — сдать тест
    """
    permission_classes = (IsStudent,)

    def post(self, request, attempt_id):
        attempt = get_object_or_404(
            QuizAttempt,
            pk=attempt_id,
            student=request.user,
            status=QuizAttempt.Status.IN_PROGRESS
        )

        now = timezone.now()
        quiz = attempt.quiz

        # Проверяем не истекло ли время
        if quiz.time_limit:
            from datetime import timedelta
            deadline = attempt.started_at + timedelta(minutes=quiz.time_limit)
            if now > deadline:
                attempt.status = QuizAttempt.Status.TIMED_OUT
                attempt.finished_at = deadline
                attempt.score = 0.0
                attempt.save()
                return Response(
                    {'detail': 'Время на выполнение теста истекло.', 'status': 'timed_out'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        serializer = SubmitQuizSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        attempt.answers = serializer.validated_data['answers']
        attempt.status = QuizAttempt.Status.COMPLETED
        attempt.finished_at = now
        attempt.score = attempt.calculate_score()
        attempt.save()

        return Response(QuizAttemptSerializer(attempt).data)


class MyQuizAttemptsView(generics.ListAPIView):
    """
    GET /api/quizzes/<quiz_id>/my-attempts/ — мои попытки по тесту
    """
    serializer_class = QuizAttemptSerializer
    permission_classes = (IsStudent,)

    def get_queryset(self):
        return QuizAttempt.objects.filter(
            quiz_id=self.kwargs['quiz_id'],
            student=self.request.user
        )


class QuizAttemptsView(generics.ListAPIView):
    """
    GET /api/quizzes/<quiz_id>/attempts/ — все попытки (teacher, admin)
    """
    serializer_class = QuizAttemptSerializer
    permission_classes = (IsTeacherOrAdmin,)

    def get_queryset(self):
        return QuizAttempt.objects.filter(
            quiz_id=self.kwargs['quiz_id']
        ).select_related('student', 'quiz')


class CourseQuizResultsView(APIView):
    """
    GET /api/courses/<course_id>/quiz-results/ — результаты тестов по курсу (teacher, admin)
    """
    permission_classes = (IsTeacherOrAdmin,)

    def get(self, request, course_id):
        from apps.courses.models import Course
        get_object_or_404(Course, pk=course_id)

        attempts = QuizAttempt.objects.filter(
            status=QuizAttempt.Status.COMPLETED,
            quiz__lesson__course_id=course_id,
        ).select_related('student', 'quiz', 'quiz__lesson')

        data = {}
        for a in attempts:
            key = (a.student_id, a.quiz_id)
            score = a.score or 0
            if key not in data:
                data[key] = {
                    'student_id': a.student_id,
                    'student_name': a.student.full_name,
                    'email': a.student.email,
                    'lesson_title': a.quiz.lesson.title,
                    'quiz_id': a.quiz_id,
                    'quiz_title': a.quiz.title,
                    'best_score': score,
                    'passing_score': a.quiz.passing_score,
                    'passed': score >= a.quiz.passing_score,
                    'attempt_count': 1,
                    'last_attempt': a.finished_at,
                }
            else:
                data[key]['attempt_count'] += 1
                if score > data[key]['best_score']:
                    data[key]['best_score'] = score
                    data[key]['passed'] = score >= data[key]['passing_score']
                if a.finished_at and (
                    data[key]['last_attempt'] is None
                    or a.finished_at > data[key]['last_attempt']
                ):
                    data[key]['last_attempt'] = a.finished_at

        results = sorted(data.values(), key=lambda x: (x['student_name'], x['quiz_title']))
        for r in results:
            if r['last_attempt']:
                r['last_attempt'] = r['last_attempt'].strftime('%d.%m.%Y %H:%M')

        return Response(results)


def _build_quiz_performance_data():
    """Returns list of dicts with best quiz score per (student, quiz)."""
    attempts = QuizAttempt.objects.filter(
        status=QuizAttempt.Status.COMPLETED
    ).select_related('student', 'quiz', 'quiz__lesson', 'quiz__lesson__course')

    data = {}
    for a in attempts:
        key = (a.student_id, a.quiz_id)
        score = a.score or 0
        if key not in data:
            data[key] = {
                'student_id': a.student_id,
                'student_name': a.student.full_name,
                'email': a.student.email,
                'course': a.quiz.lesson.course.title,
                'lesson': a.quiz.lesson.title,
                'quiz': a.quiz.title,
                'best_score': score,
                'passing_score': a.quiz.passing_score,
                'count': 1,
                'last_attempt': a.finished_at,
            }
        else:
            data[key]['count'] += 1
            if score > data[key]['best_score']:
                data[key]['best_score'] = score
            if a.finished_at and (data[key]['last_attempt'] is None or a.finished_at > data[key]['last_attempt']):
                data[key]['last_attempt'] = a.finished_at

    return list(data.values())


HEADERS = ['ID студента', 'Студент', 'Email', 'Курс', 'Урок', 'Тест', 'Лучший результат (%)', 'Пройден', 'Кол-во попыток', 'Дата последней попытки']


def _row(d):
    return [
        d['student_id'],
        d['student_name'],
        d['email'],
        d['course'],
        d['lesson'],
        d['quiz'],
        d['best_score'],
        'Да' if d['best_score'] >= d['passing_score'] else 'Нет',
        d['count'],
        d['last_attempt'].strftime('%d.%m.%Y %H:%M') if d['last_attempt'] else '',
    ]


class ExportQuizPerformanceView(APIView):
    """
    GET /api/quizzes/export/performance/ — CSV
    """
    permission_classes = (IsTeacherOrAdmin,)

    def get(self, request):
        rows = _build_quiz_performance_data()
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(HEADERS)
        for d in rows:
            writer.writerow(_row(d))
        # utf-8-sig BOM so Excel opens Cyrillic correctly
        response = HttpResponse('﻿' + output.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="quiz_performance.csv"'
        return response


class ExportQuizPerformanceExcelView(APIView):
    """
    GET /api/quizzes/export/performance/excel/ — XLSX
    """
    permission_classes = (IsTeacherOrAdmin,)

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        rows = _build_quiz_performance_data()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Успеваемость'

        # Header row styling
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)

        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for row_idx, d in enumerate(rows, start=2):
            for col_idx, value in enumerate(_row(d), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='center')
                # Colour "Пройден" column (index 7)
                if col_idx == 7:
                    if value == 'Да':
                        cell.font = Font(color='16A34A', bold=True)
                    else:
                        cell.font = Font(color='DC2626', bold=True)

        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="quiz_performance.xlsx"'
        return response